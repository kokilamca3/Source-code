from selenium import webdriver
import openpyxl

class Base1:
    def browser_launch(self):
        self.driver = webdriver.Chrome(
            executable_path=r"C:\Users\Airtel\PycharmProjects\python day9\task\chromedriver.exe")
        return self.driver

    def url_launch(self, url):
        self.driver.get(url)

    def max_window(self):
        self.driver.maximize_window()

    def user_input(self, element1, data):
        element1.send_keys(data)

    def btn_click(self, element2):
        element2.click()

    def get_data_excel(self,row_no,cell_no,element):
        excel_loc="hello.xlsx"
        w=openpyxl.Workbook()
        sheet=w.create_sheet("test",0)
        c=sheet.cell(row_no,cell_no)
        c.value=element
        w.save(excel_loc)


